# copy_sheets.py
"""
엑셀 시트 자동화 처리 스크립트
- 대상 시트 복사 및 리네임
- 동일 시트 내 값 복사 (B→E, H→I)
- '정반,xy stage-가진 안함' 시트에서
    · '정반-가진 x/y/z축'으로 값 복사 (B→B, H→H)
    · 'xy stage-가진 x/y/z축'으로 값 복사 (E→B, I→H)
- 시각화
    · '정반 가진 데이터 그림' 시트(정반-가진 x/y/z축) 그래프 18개
    · 'XY 가진 데이터 그림' 시트(xy stage-가진 x/y/z축) 그래프 18개
- (추가) 두 시트 모든 차트에 디자인 서식 적용(축/제목/눈금선)

※ 기존 Phase A~F 로직은 동일하며, 이후 시각화(Phase G) → 서식(Phase H) → 저장(Phase I)
"""

from pathlib import Path
from typing import List, Tuple, Iterable
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries, column_index_from_string
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties


# =============================================================================
# 유틸
# =============================================================================
def unique_title(wb: Workbook, desired: str) -> str:
    """시트 이름이 중복되지 않도록 고유한 이름을 반환."""
    existing = set(wb.sheetnames)
    if desired not in existing:
        return desired
    i = 1
    while True:
        cand = f"{desired} ({i})"
        if cand not in existing:
            return cand
        i += 1


def _to_float_if_numeric(val):
    """문자열 숫자 → float 변환(실패 시 원본 반환)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = val.strip()
        try:
            if v.isdigit() or (v.replace(".", "", 1).replace("-", "", 1).isdigit()):
                return float(v)
        except Exception:
            pass
    return val


def _safe_iter_sheetnames(wb: Workbook, names: Iterable[str]) -> Iterable[str]:
    """워크북에 실제 존재하는 시트명만 반환."""
    for n in names:
        if n in wb.sheetnames:
            yield n


def paste_values_intra(ws: Worksheet, src_range: str, dst_col_letter: str, dst_row_start: int) -> None:
    """
    [동일 시트] src_range를 읽어 (dst_col_letter, dst_row_start) 좌상단에 '값'만 붙여넣기.
    예: "B11:D2058" -> "E11:G2058"
    """
    min_col, min_row, max_col, max_row = range_boundaries(src_range)
    rows = max_row - min_row + 1
    cols = max_col - min_col + 1
    dst_col_start = column_index_from_string(dst_col_letter)

    for r in range(rows):
        for c in range(cols):
            src_cell = ws.cell(row=min_row + r, column=min_col + c)
            dst_cell = ws.cell(row=dst_row_start + r, column=dst_col_start + c)
            dst_cell.value = _to_float_if_numeric(src_cell.value)


def paste_values_inter(src_ws: Worksheet, src_range: str, dst_ws: Worksheet, dst_col_letter: str, dst_row_start: int) -> None:
    """
    [시트 간] src_ws의 src_range를 읽어 dst_ws의 (dst_col_letter, dst_row_start) 좌상단에 '값'만 붙여넣기.
    """
    min_col, min_row, max_col, max_row = range_boundaries(src_range)
    rows = max_row - min_row + 1
    cols = max_col - min_col + 1
    dst_col_start = column_index_from_string(dst_col_letter)

    for r in range(rows):
        for c in range(cols):
            src_cell = src_ws.cell(row=min_row + r, column=min_col + c)
            dst_cell = dst_ws.cell(row=dst_row_start + r, column=dst_col_start + c)
            dst_cell.value = _to_float_if_numeric(src_cell.value)


def bulk_copy_from_sheet(
    wb: Workbook,
    src_ws: Worksheet,
    jobs: list,
) -> None:
    """
    한 소스 시트에서 여러 전송 작업을 일괄 수행.
    jobs: 각 항목은 {"src_range", "dst_names", "dst_col_letter", "dst_row_start"} 포함
    """
    for job in jobs:
        src_range = job["src_range"]
        dst_names = job["dst_names"]
        dst_col_letter = job["dst_col_letter"]
        dst_row_start = job["dst_row_start"]
        for dst_name in _safe_iter_sheetnames(wb, dst_names):
            dst_ws = wb[dst_name]
            paste_values_inter(src_ws, src_range, dst_ws, dst_col_letter, dst_row_start)


# =============================================================================
# Phase G(시각화): ‘정반 가진 데이터 그림’ 시트에 분산형(곡선) 차트 생성
# =============================================================================
def add_ground_chart_sheet(wb):
    """
    세 소스 시트의 데이터를 사용해 '정반 가진 데이터 그림' 시트에
    '곡선이 있는 분산형 차트' 총 18개를 생성한다.

    데이터/배치:
    [정반-가진 x축]  A1,  A15,  A31,  I1,  I15,  I31
    [정반-가진 y축]  Q1,  Q15,  Q31,  Y1,  Y15,  Y31
    [정반-가진 z축]  AG1, AG15, AG31, AO1, AO15, AO31
    (각 6개: 전/후 2계열 3개 + 단일 1계열 3개)
    """
    dst_name = "정반 가진 데이터 그림"

    # 대상 시트 준비(기존 차트 삭제 후 재생성)
    if dst_name in wb.sheetnames:
        ws_dst = wb[dst_name]
        for ch in list(getattr(ws_dst, "_charts", [])):
            ws_dst._charts.remove(ch)
    else:
        ws_dst = wb.create_sheet(dst_name)

    # 공통 참조 범위
    r_min, r_max = 12, 2058
    x_col_letter = "K"
    x_col = column_index_from_string(x_col_letter)

    # 내부 유틸: 다계열(전/후) 차트
    def _add_scatter_multi(ws_src, dst_anchor: str, y_cols: list, titles: list):
        chart = ScatterChart()
        chart.varyColors = False
        chart.legend = Legend()
        chart.legend.position = "r"
        chart.title = None
        chart.x_axis.title = None
        chart.y_axis.title = None

        xref = Reference(ws_src, min_col=x_col, min_row=r_min, max_col=x_col, max_row=r_max)

        for col_letter, title in zip(y_cols, titles):
            cY = column_index_from_string(col_letter)
            yref = Reference(ws_src, min_col=cY, min_row=r_min, max_col=cY, max_row=r_max)
            s = Series(yref, xref)
            s.title = SeriesLabel(v=title)  # 문자열 제목 강제
            s.smooth = True
            if hasattr(s, "marker") and s.marker:
                s.marker.symbol = "none"
            chart.series.append(s)

        ws_dst.add_chart(chart, dst_anchor)

    # 내부 유틸: 단일 계열 차트(정반 가진 전/후)
    def _add_scatter_single(ws_src, dst_anchor: str, y_col: str, title: str = "정반 가진 전/후"):
        chart = ScatterChart()
        chart.varyColors = False            # 포인트별 색상 금지 → 범례 1개 유지
        chart.legend = Legend()
        chart.legend.position = "r"
        chart.title = None
        chart.x_axis.title = None
        chart.y_axis.title = None

        xref = Reference(ws_src, min_col=x_col, min_row=r_min, max_col=x_col, max_row=r_max)
        cY = column_index_from_string(y_col)
        yref = Reference(ws_src, min_col=cY, min_row=r_min, max_col=cY, max_row=r_max)

        s = Series(yref, xref)
        s.title = SeriesLabel(v=title)
        s.smooth = True
        if hasattr(s, "marker") and s.marker:
            s.marker.symbol = "none"

        chart.series.append(s)
        ws_dst.add_chart(chart, dst_anchor)

    # 시트별 작업 정의
    jobs = [
        # (소스시트명, [(앵커, 타입, ycols/titles or ycol)])
        ("정반-가진 x축", [
            ("A1",  "multi", {"y_cols": ["L", "M"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("A15", "multi", {"y_cols": ["N", "O"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("A31", "multi", {"y_cols": ["R", "S"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("I1",  "single", {"y_col": "U"}),
            ("I15", "single", {"y_col": "V"}),
            ("I31", "single", {"y_col": "W"}),
        ]),
        ("정반-가진 y축", [
            ("Q1",  "multi", {"y_cols": ["L", "M"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("Q15", "multi", {"y_cols": ["N", "O"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("Q31", "multi", {"y_cols": ["R", "S"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("Y1",  "single", {"y_col": "U"}),
            ("Y15", "single", {"y_col": "V"}),
            ("Y31", "single", {"y_col": "W"}),
        ]),
        ("정반-가진 z축", [
            ("AG1",  "multi", {"y_cols": ["L", "M"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("AG15", "multi", {"y_cols": ["N", "O"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("AG31", "multi", {"y_cols": ["R", "S"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("AO1",  "single", {"y_col": "U"}),
            ("AO15", "single", {"y_col": "V"}),
            ("AO31", "single", {"y_col": "W"}),
        ]),
    ]

    # 실행
    for src_name, task_list in jobs:
        if src_name not in wb.sheetnames:
            continue
        ws_src = wb[src_name]
        for anchor, t, payload in task_list:
            if t == "multi":
                _add_scatter_multi(ws_src, anchor, payload["y_cols"], payload["titles"])
            else:
                _add_scatter_single(ws_src, anchor, payload["y_col"])


def add_xy_stage_chart_sheet(wb):
    """
    소스: 'xy stage-가진 x축', 'xy stage-가진 y축', 'xy stage-가진 z축'
    대상: 'XY 가진 데이터 그림' 시트
    내용: '곡선이 있는 분산형' 차트 18개 생성 (배치/데이터는 정반-가진 시각화와 동일)
    """
    dst_name = "XY 가진 데이터 그림"

    # 대상 시트 준비(기존 차트 삭제 후 재생성)
    if dst_name in wb.sheetnames:
        ws_dst = wb[dst_name]
        for ch in list(getattr(ws_dst, "_charts", [])):
            ws_dst._charts.remove(ch)
    else:
        ws_dst = wb.create_sheet(dst_name)

    # 공통 참조 범위
    r_min, r_max = 12, 2058
    x_col_letter = "K"
    x_col = column_index_from_string(x_col_letter)

    # 내부 유틸: 다계열(전/후) 차트
    def _add_scatter_multi(ws_src, dst_anchor: str, y_cols: list, titles: list):
        chart = ScatterChart()
        chart.varyColors = False
        chart.legend = Legend()
        chart.legend.position = "r"
        chart.title = None
        chart.x_axis.title = None
        chart.y_axis.title = None

        xref = Reference(ws_src, min_col=x_col, min_row=r_min, max_col=x_col, max_row=r_max)

        for col_letter, title in zip(y_cols, titles):
            cY = column_index_from_string(col_letter)
            yref = Reference(ws_src, min_col=cY, min_row=r_min, max_col=cY, max_row=r_max)
            s = Series(yref, xref)
            s.title = SeriesLabel(v=title)
            s.smooth = True
            if hasattr(s, "marker") and s.marker:
                s.marker.symbol = "none"
            chart.series.append(s)

        ws_dst.add_chart(chart, dst_anchor)

    # 내부 유틸: 단일 계열 차트(정반 가진 전/후)
    def _add_scatter_single(ws_src, dst_anchor: str, y_col: str, title: str = "정반 가진 전/후"):
        chart = ScatterChart()
        chart.varyColors = False
        chart.legend = Legend()
        chart.legend.position = "r"
        chart.title = None
        chart.x_axis.title = None
        chart.y_axis.title = None

        xref = Reference(ws_src, min_col=x_col, min_row=r_min, max_col=x_col, max_row=r_max)
        cY = column_index_from_string(y_col)
        yref = Reference(ws_src, min_col=cY, min_row=r_min, max_col=cY, max_row=r_max)

        s = Series(yref, xref)
        s.title = SeriesLabel(v=title)
        s.smooth = True
        if hasattr(s, "marker") and s.marker:
            s.marker.symbol = "none"

        chart.series.append(s)
        ws_dst.add_chart(chart, dst_anchor)

    # 소스 시트별 작업 정의 (xy stage-가진 *)
    jobs = [
        ("xy stage-가진 x축", [
            ("A1",  "multi", {"y_cols": ["L", "M"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("A15", "multi", {"y_cols": ["N", "O"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("A31", "multi", {"y_cols": ["R", "S"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("I1",  "single", {"y_col": "U"}),
            ("I15", "single", {"y_col": "V"}),
            ("I31", "single", {"y_col": "W"}),
        ]),
        ("xy stage-가진 y축", [
            ("Q1",  "multi", {"y_cols": ["L", "M"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("Q15", "multi", {"y_cols": ["N", "O"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("Q31", "multi", {"y_cols": ["R", "S"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("Y1",  "single", {"y_col": "U"}),
            ("Y15", "single", {"y_col": "V"}),
            ("Y31", "single", {"y_col": "W"}),
        ]),
        ("xy stage-가진 z축", [
            ("AG1",  "multi", {"y_cols": ["L", "M"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("AG15", "multi", {"y_cols": ["N", "O"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("AG31", "multi", {"y_cols": ["R", "S"], "titles": ["정반 가진 전", "정반 가진 후"]}),
            ("AO1",  "single", {"y_col": "U"}),
            ("AO15", "single", {"y_col": "V"}),
            ("AO31", "single", {"y_col": "W"}),
        ]),
    ]

    # 실행
    for src_name, task_list in jobs:
        if src_name not in wb.sheetnames:
            continue
        ws_src = wb[src_name]
        for anchor, t, payload in task_list:
            if t == "multi":
                _add_scatter_multi(ws_src, anchor, payload["y_cols"], payload["titles"])
            else:
                _add_scatter_single(ws_src, anchor, payload["y_col"])


# =============================================================================
# Phase H. 차트 서식 일괄 적용
#  - 대상 시트: "정반 가진 데이터 그림", "XY 가진 데이터 그림"
#  - 모든 차트에 적용
#    (1) 축: 가로 0~300 / majorUnit=50, 세로 자동
#    (2) 축 제목: X= "Frequency (Hz)"(Bold), Y= (단일계열) "Ratio", (2계열) "Power Spectrum (μG/√Hz)"
#    (3) 눈금선: 기본 주 가로/세로, 색상 #BFBFBF
# =============================================================================
def _set_axis_title(axis, text, bold=False):
    axis.title = text
    if not bold:
        return
    try:
        if axis.title and axis.title.tx and axis.title.tx.rich and axis.title.tx.rich.p:
            for p in axis.title.tx.rich.p:
                if getattr(p, "r", None):
                    for r in p.r:
                        from openpyxl.drawing.text import CharacterProperties
                        if r.rPr is None:
                            r.rPr = CharacterProperties()
                        r.rPr.b = True
    except Exception:
        pass

def _set_gridlines(axis, major_color="BFBFBF", minor_color="E6E6E6"):
    """주/보조 눈금선 색 지정 및 활성화"""
    # Major
    ml = ChartLines()
    ml.spPr = GraphicalProperties(ln=LineProperties(solidFill=major_color))
    axis.majorGridlines = ml
    # Minor
    try:
        mnl = ChartLines()
        mnl.spPr = GraphicalProperties(ln=LineProperties(solidFill=minor_color))
        axis.minorGridlines = mnl
    except Exception:
        # 일부 축 타입/버전에서 minorGridlines 미지원 → 무시
        pass


def style_charts_for_sheets(wb, sheet_names=("정반 가진 데이터 그림", "XY 가진 데이터 그림")):
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        charts = list(getattr(ws, "_charts", []))
        if not charts:
            continue

        for ch in charts:
            # ── X축 범위/간격 고정
            try:
                ch.x_axis.scaling.min = 0
                ch.x_axis.scaling.max = 300
                ch.x_axis.majorUnit = 50
                try:
                    ch.x_axis.minorUnit = 50
                except Exception:
                    pass
            except Exception:
                pass

            # ── Y축 자동 범위
            try:
                try: ch.y_axis.scaling.min = None
                except Exception: pass
                try: ch.y_axis.scaling.max = None
                except Exception: pass
            except Exception:
                pass

            # ▼▼ 축 "라벨이 보이도록" 강제 설정(핵심) ▼▼
            try:
                # 축을 숨기지 않도록
                ch.x_axis.delete = False
                ch.y_axis.delete = False

                # 라벨 위치 (값 표시)
                ch.x_axis.tickLblPos = "nextTo"
                ch.y_axis.tickLblPos = "nextTo"

                # 축 교차: 0 근방에서 자동 (라벨 렌더링 보장에 유리)
                ch.x_axis.crosses = "autoZero"
                ch.y_axis.crosses = "autoZero"

                # 눈금 표시(라벨과 함께 보통 'out'이 깔끔)
                ch.x_axis.majorTickMark = "out"
                ch.y_axis.majorTickMark = "out"
                ch.x_axis.minorTickMark = "out"   # 보조 눈금 표시(라벨은 메이저만)
                ch.y_axis.minorTickMark = "out"

                # 숫자 포맷(라벨 간결)
                ch.x_axis.number_format = "0"     # 0, 50, 100 ...
                # 필요 시 Y축 포맷: ch.y_axis.number_format = "0.000"
            except Exception:
                pass
            # ▲▲ 여기까지가 라벨 강제 노출 핵심 ▲▲

            # 축 제목
            _set_axis_title(ch.x_axis, "Frequency (Hz)", bold=True)
            y_title = "Ratio" if len(getattr(ch, "series", [])) == 1 else "Power Spectrum (μG/√Hz)"
            _set_axis_title(ch.y_axis, y_title, bold=False)

            # 그리드(주/보조)
            _set_gridlines(ch.x_axis, major_color="BFBFBF", minor_color="E6E6E6")
            _set_gridlines(ch.y_axis, major_color="BFBFBF", minor_color="E6E6E6")

            # (선택) 차트 크기 넉넉히 (라벨 겹침 방지)
            try:
                # 현재 크기 대비 약 85% 수준으로 축소 (여백 확보)
                ch.width = getattr(ch, "width", 6.5) * 0.95
                ch.height = getattr(ch, "height", 4.5) * 0.95

                # 추가로 축 제목이 아래/왼쪽으로 너무 붙을 경우, 약간 이동 여유
                if hasattr(ch, "x_axis"):
                    ch.x_axis.tickLblPos = "nextTo"
                    ch.x_axis.crosses = "autoZero"
                if hasattr(ch, "y_axis"):
                    ch.y_axis.tickLblPos = "nextTo"
                    ch.y_axis.crosses = "autoZero"
            except Exception:
                pass



            # ── 축 제목
            _set_axis_title(ch.x_axis, "Frequency (Hz)", bold=True)
            y_title = "Ratio" if len(getattr(ch, "series", [])) == 1 else "Power Spectrum (μG/√Hz)"
            _set_axis_title(ch.y_axis, y_title, bold=False)

            # ── 눈금선: 주(#BFBFBF) + 보조(#E6E6E6)
            _set_gridlines(ch.x_axis, major_color="BFBFBF", minor_color="E6E6E6")
            _set_gridlines(ch.y_axis, major_color="BFBFBF", minor_color="E6E6E6")
        


# =============================================================================
# MAIN
# =============================================================================
def copy_target_sheets(src_path: Path) -> Tuple[Path, List[str]]:
    """
    작업 순서(Phase):
      A. 파일 검증 및 로드
      B. 대상 시트 복사 → 복사본 리네임(NEW_COPY_NAMES)
      C. 원본 시트 리네임(NEW_ORIGINAL_NAMES)
      D. (원본 각 시트) 동일 시트 내 값 복사
      E+F. '정반,xy stage-가진 안함'에서 다수 대상 시트로 값 복사(통합)
      G. 시각화: 두 시트에 차트 생성
      H. 서식: 두 시트 차트 서식 일괄 적용
      I. 저장 및 결과 반환
    """
    if not src_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {src_path}")

    # -------------------------------------------------------------------------
    # Phase A. 로드
    # -------------------------------------------------------------------------
    if src_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError("지원하지 않는 파일 형식입니다. (.xlsx, .xlsm 만 허용)")
    wb = load_workbook(src_path, data_only=False)  # 수식 유지

    # -------------------------------------------------------------------------
    # 대상/새 이름 정의
    # -------------------------------------------------------------------------
    TARGET_NAMES = [
        "정반,xy stage-가진 x축(3)",
        "정반,xy stage-가진 y축(3)",
        "정반,xy stage-가진 z축(3)",
    ]
    NEW_COPY_NAMES = [  # 복사본(값 처리 대상)
        "xy stage-가진 x축",
        "xy stage-가진 y축",
        "xy stage-가진 z축",
    ]
    NEW_ORIGINAL_NAMES = [  # 원본 리네임(정반-가진 *)
        "정반-가진 x축",
        "정반-가진 y축",
        "정반-가진 z축",
    ]
    SRC_NO_DRIVE = "정반,xy stage-가진 안함"

    # -------------------------------------------------------------------------
    # Phase B. 대상 시트 복사 → 복사본 리네임
    # -------------------------------------------------------------------------
    created_names: List[str] = []
    for src_name, new_copy in zip(TARGET_NAMES, NEW_COPY_NAMES):
        if src_name not in wb.sheetnames:
            raise KeyError(f"대상 시트가 없습니다: {src_name}")
        src_ws = wb[src_name]
        new_ws = wb.copy_worksheet(src_ws)
        new_ws.title = unique_title(wb, new_copy)
        created_names.append(new_ws.title)

    # -------------------------------------------------------------------------
    # Phase C. 원본 시트 리네임(정반-가진 *)
    # -------------------------------------------------------------------------
    for src_name, new_origin in zip(TARGET_NAMES, NEW_ORIGINAL_NAMES):
        if src_name not in wb.sheetnames:
            continue
        ws = wb[src_name]
        ws.title = unique_title(wb, new_origin)

    # -------------------------------------------------------------------------
    # Phase D. (원본 각 시트) 동일 시트 내 값 복사
    #   B11:D2058 → E11:G2058
    #   H11:H2058 → I11:I2058
    # -------------------------------------------------------------------------
    for name in _safe_iter_sheetnames(wb, NEW_ORIGINAL_NAMES):
        ws = wb[name]
        paste_values_intra(ws, "B11:D2058", "E", 11)
        paste_values_intra(ws, "H11:H2058", "I", 11)

    # -------------------------------------------------------------------------
    # Phase E+F. '정반,xy stage-가진 안함'에서 다수 대상 시트로 값 복사(통합)
    # 원본(정반-가진 *):   B→B, H→H
    # 복사본(xy stage-*) : E→B, I→H
    # -------------------------------------------------------------------------
    if SRC_NO_DRIVE not in wb.sheetnames:
        raise KeyError(f"필요 시트가 없습니다: {SRC_NO_DRIVE}")
    src_ws = wb[SRC_NO_DRIVE]

    bulk_copy_from_sheet(
        wb,
        src_ws,
        jobs=[
            # → 원본(정반-가진 x/y/z축)
            {"src_range": "B11:D2058", "dst_names": NEW_ORIGINAL_NAMES, "dst_col_letter": "B", "dst_row_start": 11},
            {"src_range": "H11:H2058", "dst_names": NEW_ORIGINAL_NAMES, "dst_col_letter": "H", "dst_row_start": 11},
            # → 복사본(xy stage-가진 x/y/z축)
            {"src_range": "E11:G2058", "dst_names": NEW_COPY_NAMES,     "dst_col_letter": "B", "dst_row_start": 11},
            {"src_range": "I11:I2058", "dst_names": NEW_COPY_NAMES,     "dst_col_letter": "H", "dst_row_start": 11},
        ],
    )

    # -------------------------------------------------------------------------
    # Phase G. 시각화 생성
    # -------------------------------------------------------------------------
    add_ground_chart_sheet(wb)      # 정반-가진 x/y/z 그래프 18개
    add_xy_stage_chart_sheet(wb)    # xy stage-가진 x/y/z 그래프 18개

    # -------------------------------------------------------------------------
    # Phase H. 차트 서식 적용 (축/제목/눈금선)
    # -------------------------------------------------------------------------
    style_charts_for_sheets(wb, ("정반 가진 데이터 그림", "XY 가진 데이터 그림"))

    # -------------------------------------------------------------------------
    # Phase I. 저장 및 결과 반환
    # -------------------------------------------------------------------------
    dst_path = src_path.with_name(f"{src_path.stem}_operated{src_path.suffix}")
    wb.save(dst_path)
    return dst_path, created_names


# =============================================================================
# CLI (예시)
# =============================================================================
if __name__ == "__main__":
    example = r"C:\Users\JaeSeon.Yu\vscode\Company\2025.10.01 XY 스테이지 더미 측정.xlsx"
    try:
        new_file, copied_names = copy_target_sheets(Path(example))
        print("생성 완료 ✅")
        print(f"저장 파일: {new_file}")
        print("복사된 시트:")
        for n in copied_names:
            print(f" - {n}")
    except Exception as e:
        print(f"오류 발생 ❌: {e}")
