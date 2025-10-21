from pathlib import Path
from typing import List, Union
from openpyxl import load_workbook, Workbook

# ===== 사용자 설정 =====
INPUT_FILE = r"C:\path\to\원본.xlsx"  # 입력 파일 경로
OUTPUT_FILE = "완성본.xlsx"           # 출력 파일명
TARGET_SHEETS = ["1", "2", "3", "4", "5"]
TARGET_CELL = "B11"
OUTPUT_SHEET_NAME = "6"

Number = Union[int, float]

def read_b11_values(filepath: str, sheets: List[str], cell: str) -> List[Number]:
    """입력 엑셀에서 지정된 시트들의 cell 값을 숫자로 읽어 리스트로 반환"""
    p = Path(filepath)
    if not p.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {filepath}")

    try:
        wb = load_workbook(filepath, data_only=True)  # 수식 결과값이 있으면 값으로 읽음
    except Exception as e:
        raise RuntimeError(f"엑셀 파일을 열 수 없습니다: {e}")

    values: List[Number] = []
    for name in sheets:
        if name not in wb.sheetnames:
            raise ValueError(f"대상 시트가 없습니다: '{name}'")
        ws = wb[name]
        val = ws[cell].value

        # 유효성 검사: 숫자형만 허용
        if val is None or not isinstance(val, (int, float)):
            raise TypeError(f"시트 '{name}'의 {cell} 값이 숫자가 아닙니다: {val!r}")

        values.append(val)

    if len(values) != len(sheets):
        raise RuntimeError("값 수집 개수가 시트 개수와 일치하지 않습니다.")

    return values

def compute_average(values: List[Number]) -> float:
    """숫자 리스트의 산술평균 계산"""
    if not values:
        raise ValueError("평균 계산을 위한 값 리스트가 비어있습니다.")
    return float(sum(values) / len(values))

def build_output_workbook(avg_value: float, sheet_name: str, cell: str) -> Workbook:
    """새 워크북을 만들고 지정 시트/셀에 평균값을 기록하여 반환"""
    new_wb = Workbook()
    ws = new_wb.active
    ws.title = sheet_name
    ws[cell].value = avg_value
    # 필요 시 표시 형식 지정: ws[cell].number_format = "0.0000"
    return new_wb

def save_workbook(wb: Workbook, output_path: str) -> None:
    """워크북 저장"""
    try:
        wb.save(output_path)
    except Exception as e:
        raise RuntimeError(f"파일 저장 실패: {e}")

def main():
    print("[1/4] 입력 파일에서 값 읽기...")
    values = read_b11_values(INPUT_FILE, TARGET_SHEETS, TARGET_CELL)
    print(f"    - 읽은 값: {values}")

    print("[2/4] 평균 계산...")
    avg = compute_average(values)
    print(f"    - 평균값: {avg}")

    print("[3/4] 출력 워크북 구성(시트 '6'만 존재)...")
    out_wb = build_output_workbook(avg, OUTPUT_SHEET_NAME, TARGET_CELL)

    print("[4/4] 파일 저장...")
    save_workbook(out_wb, OUTPUT_FILE)
    print(f"완료: '{OUTPUT_FILE}' 에 시트 '{OUTPUT_SHEET_NAME}'의 {TARGET_CELL} = {avg} 저장")

if __name__ == "__main__":
    main()
